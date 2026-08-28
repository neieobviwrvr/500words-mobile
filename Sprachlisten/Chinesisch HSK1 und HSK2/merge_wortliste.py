# -*- coding: utf-8 -*-
"""
Fuehrt Simons Excel-Wortliste und die ergaenzten Woerter zu EINER Datei
zusammen und schreibt das Ergebnis des Quellen-Abgleichs mit hinein.

Aufruf:  python merge_wortliste.py
Ergebnis: "Chinesische Wortliste (ergaenzt).xlsx" - Simons Originaldatei
bleibt unangetastet.

ABGLEICH (2026-08-20)
---------------------
Geprueft gegen zwei voneinander unabhaengige, maschinenlesbare
Neuveroeffentlichungen der HSK-2.0-Wortliste:

  Quelle A  github.com/plaktos/hsk_csv            (hsk1.csv + hsk2.csv)
  Quelle B  github.com/drkameleon/complete-hsk-vocabulary
                                                  (wordlists/exclusive/old)

Beide liefern 297 eindeutige Woerter fuer HSK 1 + 2 und stimmen in ALLEN
297 ueberein - null Abweichung zwischen den Quellen. Das ist der Grund,
warum ihr Ergebnis hier als belastbar behandelt wird.

Keine der beiden ist die Primaerquelle; die waere der Pruefungsleitfaden
von Chinese Testing International. Beide geben denselben standardisierten
Wortschatz wieder.

WAS DER ABGLEICH ZEIGT
----------------------
Von den 311 Eintraegen sind 215 in HSK 1+2 belegt, 96 nicht. Das ist KEIN
Fehler, sondern eine Aussage ueber die Stufe:

  * Viele davon sind HSK 3 (aengste, 老 statt 老师; 手 statt 手机) oder
    stehen in der Liste nur als laengeres Wort (说话 statt 说,
    分钟 statt 分).
  * Andere sind bewusst gewaehlt und stehen in keinem Lehrbuch:
    微信, 酒店, 地铁站, 单身 - genau die Woerter, die Backpacker und
    Austauschstudenten brauchen.

Die HSK-Stufe der nicht belegten Eintraege wird deshalb auf
"nicht in HSK 1+2" gesetzt statt geraten. Die Woerter selbst bleiben drin.
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

import build_chinesisch_kurs as bk
import chinesisch_erweiterung as erw

HIER = Path(__file__).parent
ZIEL = HIER / "Chinesische Wortliste (ergaenzt).xlsx"

# Ergebnis des Abgleichs: Eintraege, die WEDER Quelle A NOCH Quelle B in
# HSK 1+2 fuehrt. Zum Nachpruefen die beiden Quellen oben erneut ziehen.
NICHT_IN_HSK12 = set("""
说 关 用 短 分 里面 外面 上面 下面 面条 面包 饿 渴 吃饱 胖 瘦 老 年轻 聪明
可爱 蓝 应该 看书 上网 音乐 英语 国家 上海 德国 以前 以后 第一次 手 耳朵
鼻子 嘴 舒服 下雪 晴天 阴天 爱好 旅行 你好 礼物 一下 帮 不去 今天晚上
喝一杯 单身 同事 男朋友 女朋友 帅 有趣 微信 等 吃饭 什么时候 回家 哪里
饭店 洗手间 这里 那里 酒店 地铁站 啤酒 好喝 甜 棒 菜单 更 过 拿 翻译 解释
头 肚子 腿 感冒 发烧 难受 疼 清楚 周末 打算 安排 兴趣 习惯 了解 加 聊天
逛街 帮忙 照顾
""".split())

# Welche HSK-1+2-Woerter in unserer Liste noch FEHLEN, sortiert nach
# Nutzen fuer die Zielgruppe (Backpacker/Austauschstudenten in den
# Zwanzigern). Die Einteilung ist eine EMPFEHLUNG, keine Messung - Simon
# entscheidet.
#
# Warum nicht einfach alle 82 aufnehmen: bei 4-5 neuen Vokabeln je Lektion
# waeren das rund 18 zusaetzliche Lektionen, der Kurs waechst von 92 auf
# ueber 110. Bei 10-15 Minuten taeglich ist das der Unterschied zwischen
# etwa drei und vier Monaten - und beworben werden ZWEI Monate bis zum
# Alltagsgespraech. HSK ist eine Pruefungsordnung, keine Haeufigkeitsliste.
# ERLEDIGT 2026-08-20: diese 40 sind auf Simons Entscheidung in den Kurs
# aufgenommen worden ("dann kostet es eben Lektionen") und stehen jetzt in
# chinesisch_erweiterung.py. Sie fehlen also nicht mehr - die Liste bleibt
# leer, damit Blatt 2 nur noch echte Luecken zeigt.
FEHLEND_AUFNEHMEN = []

FEHLEND_GRENZFALL = """
爱 看见 笑 出 离 每 日 去年 上班 同学 大家 孩子 男人 女人 事情 希望
第一 自行车 电视 手表 姓 为 向 得 着 它 饭馆
""".split()

FEHLEND_WEGLASSEN = """
打篮球 公斤 踢 船 雪 晴 阴 课 题 字 本 张 件 妻子 丈夫
""".split()

FEHLENDE_HSK12 = FEHLEND_AUFNEHMEN + FEHLEND_GRENZFALL + FEHLEND_WEGLASSEN


def main():
    lex, _ = bk.lade_wortliste()

    # Woher stammt der Eintrag?
    aus_erweiterung = set(erw.ZUSATZ_WORTLISTE) | set(bk.ERGAENZUNGEN)

    # In welcher Lektion kommt das Wort als Slot vor?
    lektion_von = {}
    for _, _, lektionen in bk.module_mit_erweiterung():
        for lid, art, rahmen, pron, gruppen in lektionen:
            for h in [x for g in gruppen for x in g]:
                lektion_von.setdefault(h, []).append(lid)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Wortliste"

    kopf = ["Hanzi", "Pinyin", "Bedeutung (DE)", "HSK-Level", "Lektion",
            "Herkunft", "Abgleich"]
    ws.append(kopf)
    for zelle in ws[1]:
        zelle.font = Font(bold=True)

    grau = PatternFill("solid", fgColor="FFF2E8")
    belegt = offen = 0
    for hanzi, e in lex.items():
        in_hsk = hanzi not in NICHT_IN_HSK12
        belegt += in_hsk
        offen += (not in_hsk)
        ws.append([
            hanzi,
            e["pinyin"],
            e["de"],
            e["hsk"] if in_hsk else "nicht in HSK 1+2",
            ", ".join(lektion_von.get(hanzi, [])) or "— (nur im Satzrahmen)",
            "Erweiterung" if hanzi in aus_erweiterung else "Simons Excel",
            "in beiden Quellen belegt" if in_hsk
            else "über HSK 1+2 hinaus (meist HSK 3)",
        ])
        if not in_hsk:
            for zelle in ws[ws.max_row]:
                zelle.fill = grau

    # Legende, damit die Spalte "Abgleich" ohne Rueckfrage verstaendlich ist.
    ws.append([])
    ws.append(["Legende:", "\"in beiden Quellen belegt\" = steht in der offiziellen "
               "HSK-1+2-Liste.  \"über HSK 1+2 hinaus\" = steht dort NICHT - "
               "meist HSK 3 (老, 手, 音乐), oder die Liste fuehrt nur die "
               "laengere Form (说话 statt 说), oder das Wort ist juenger als "
               "die Liste (微信). KEIN Fehler und kein Grund zum Streichen."])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    breiten = {"A": 12, "B": 18, "C": 30, "D": 18, "E": 16, "F": 14, "G": 26}
    for spalte, breite in breiten.items():
        ws.column_dimensions[spalte].width = breite
    ws.freeze_panes = "A2"

    # Zweites Blatt: was uns aus HSK 1+2 noch fehlt
    ws2 = wb.create_sheet("Fehlt aus HSK 1+2")
    ws2.append(["Hanzi", "Empfehlung", "Begruendung"])
    for zelle in ws2[1]:
        zelle.font = Font(bold=True)

    gruen = PatternFill("solid", fgColor="E8F5E9")
    gelb = PatternFill("solid", fgColor="FFF8E1")
    rot = PatternFill("solid", fgColor="FCE8E6")
    bloecke = [
        (FEHLEND_AUFNEHMEN, "aufnehmen", gruen,
         "Echte Luecke: Funktionswoerter, Grundverben, Verkehrsmittel, "
         "Zeit- und Geldangaben. Ohne die fehlt Alltagssprache."),
        (FEHLEND_GRENZFALL, "Grenzfall", gelb,
         "Brauchbar, aber nicht dringend - haengt davon ab, wie lang der "
         "Kurs werden darf."),
        (FEHLEND_WEGLASSEN, "weglassen", rot,
         "Lehrbuch-Vokabular oder Zaehlwoerter ohne Nutzen fuer die "
         "Zielgruppe. 字 ist besonders unpassend: wir lehren keine Zeichen."),
    ]
    for woerter, empfehlung, farbe, grund in bloecke:
        for i, h in enumerate(woerter):
            ws2.append([h, empfehlung, grund if i == 0 else ""])
            for zelle in ws2[ws2.max_row]:
                zelle.fill = farbe

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 70
    ws2.freeze_panes = "A2"

    ziel = ZIEL
    try:
        wb.save(ziel)
    except PermissionError:
        # Datei ist in Excel geoeffnet - dann daneben ablegen statt scheitern.
        ziel = ZIEL.with_name(ZIEL.stem + " v2.xlsx")
        wb.save(ziel)
        print("Hinweis: Originaldatei war gesperrt (in Excel geoeffnet).")
    print(f"Geschrieben: {ziel.name}")
    print(f"  {len(lex)} Eintraege gesamt")
    print(f"    {belegt} in beiden Quellen als HSK 1+2 belegt")
    print(f"    {offen} in keiner Quelle - Stufe auf 'nicht in HSK 1+2' gesetzt")
    print(f"  {len(FEHLENDE_HSK12)} HSK-1+2-Woerter fehlen uns noch (Blatt 2)")


if __name__ == "__main__":
    main()
